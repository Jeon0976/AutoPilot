import requests
import openpyxl

# Kakao API 키 설정 
KAKAO_API_KEY = "KakaoAK a29374543e1f8c5fdcc063c3cfca8c77"
URL = "https://dapi.kakao.com/v2/local/search/address"

# x, y 좌표 반환 함수 
def get_coordinates(address):
    headers = { 
        "Authorization": KAKAO_API_KEY
    }

    params = { 
        "query": address
    }

    respone = requests.get(URL, headers=headers, params=params)
    
    data = respone.json()

    if not data['documents']: 
        return None,None
    return data['documents'][0]['x'], data['documents'][0]['y']

wb = openpyxl.load_workbook("jeonju_jeonbuk_moa-2.xlsx")
ws = wb.worksheets[0]

row = 3
while ws[f"C{row}"].value:
    address = ws[f"C{row}"].value
    lat, lng = get_coordinates(address)
    if lat and lng:
        ws[f"D{row}"].value = lat
        ws[f"E{row}"].value = lng
    row += 1

wb.save("update.xlsx")
print("End Work")